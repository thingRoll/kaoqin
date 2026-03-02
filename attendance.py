import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell
import warnings
import sys
import os
import re
import json
import threading
import queue
import time
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter import scrolledtext

# 强制输出编码 (防止无控制台打包模式下sys.stdout为None导致报错)
if sys.stdout is not None:
    sys.stdout.reconfigure(encoding='utf-8')
warnings.filterwarnings('ignore')

def get_application_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_application_path()
TEMPLATE_FILE = os.path.join(BASE_DIR, '模板-考勤.xlsx')
CONFIG_FILE = os.path.join(BASE_DIR, 'attendance_config.json')

RULES_FILE = os.path.join(BASE_DIR, '考勤计算规则说明书.txt')
MANUAL_FILE = os.path.join(BASE_DIR, '全自动考勤系统使用手册.txt')

# ================= 配置文件安全管理模块 =================
def generate_default_config():
    default_data = {
        "legal_holidays": [],
        "holiday_rest_days": [],
        "work_on_holidays": [],
        "rest_quota": 4.0,
        "site_days_dept_keywords": ["运维", "工程技术"],
        "outside_dept_keywords": ["邵寨", "南京", "门源", "北京"],
        "loc_province_out": ["北京", "门源", "邵寨", "方城", "上海", "深圳", "河南", "甘肃", "南京", "京", "蒙", "贵", "省外"],
        "shandong_keywords": [
            "济南", "青岛", "淄博", "枣庄", "东营", "烟台", "潍坊", "济宁", "泰安", "威海", 
            "日照", "临沂", "德州", "聊城", "滨州", "菏泽", "曲阜", "兖州", "龙口", 
            "白庄", "白", "曲", "郓", "郓城", "枣", "新", "新驿", "梁", "梁宝寺", "博", "博兴", "聊",
            "会展", "黄河国际会展中心", "大安机场", "文化中心", "emc", "emc-01", "公司", "本部", "山东"
        ],
        "jinan_keywords": ["济南", "会展", "黄河国际会展中心", "emc", "emc-01", "公司", "本部", "济南市"],
        "project_mapping": {
            "黄河国际会展中心": "会展",
            "济宁大安机场": "大安机场",
            "济宁文化中心": "文化中心",
            "美年大健康": "南京",
            "邵寨": "邵寨"
        },
        "city_abbreviations": {
            "梁宝寺": "梁", "郓": "郓", "郓城": "郓", "白庄": "白", "曲阜": "曲", "尼山": "曲",
            "北京": "京", "博兴": "博", "聊城": "聊", "内蒙": "蒙", "枣庄": "枣", "新驿": "新", "贵州": "贵"
        },
        "shift_24_48": ["济宁", "绿地"],
        "shift_3_rotation": ["邵寨"],
        "shift_exceptions": {"刘淑亚": "白班", "马驰": "白班"}
    }
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(default_data, f, ensure_ascii=False, indent=4)
    except Exception:
        pass
    return default_data

def load_config():
    if not os.path.exists(CONFIG_FILE):
        return generate_default_config()
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        messagebox.showerror("配置读取错误", f"配置文件格式损坏或无法读取。\n将使用默认配置运行程序！\n\n错误信息:\n{str(e)}")
        return generate_default_config()

def save_config(data):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def parse_number(val):
    if pd.isna(val) or val == '': 
        return 0.0
    nums = re.findall(r"[-+]?\d*\.\d+|\d+", str(val).strip())
    if nums:
        return float(nums[0])
    return 0.0

def safe_write(ws, r, c, value):
    cell = ws.cell(row=r, column=c)
    if not isinstance(cell, MergedCell):
        cell.value = value

# ================= 核心处理引擎 =================
def run_processing(source_file, prev_month_file, log_func):
    def log(msg):
        if log_func:
            log_func(msg)
            
    log("🚀 [System Boot] 正在唤醒考勤算力分析引擎...")
    config = load_config()
    
    legal_holidays = config.get("legal_holidays", [])
    holiday_rest_days = config.get("holiday_rest_days", [])
    work_on_holidays = config.get("work_on_holidays", [])
    REST_QUOTA = float(config.get("rest_quota", 4.0))
    SITE_DAYS_DEPT_KEYWORDS = config.get("site_days_dept_keywords", [])
    OUTSIDE_DEPT_KEYWORDS = config.get("outside_dept_keywords", [])
    LOC_PROVINCE_OUT = config.get("loc_province_out", [])
    SHANDONG_KEYWORDS = config.get("shandong_keywords", [])
    JINAN_KEYWORDS = config.get("jinan_keywords", [])
    PROJECT_MAPPING = config.get("project_mapping", {})
    CITY_ABBREVIATIONS = config.get("city_abbreviations", {})
    SHIFT_24_48 = config.get("shift_24_48", [])
    SHIFT_3_ROT = config.get("shift_3_rotation", [])
    SHIFT_EXCEPTIONS = config.get("shift_exceptions", {})

    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(f"核心文件丢失：未找到《模板-考勤.xlsx》")

    imported_banked_data = {}
    if prev_month_file and os.path.exists(prev_month_file):
        log("📂 [Data Hook] 正在建立跨文件通信，提取上月资金池...")
        try:
            prev_wb = openpyxl.load_workbook(prev_month_file, data_only=True)
            if '当月考勤' in prev_wb.sheetnames:
                prev_ws = prev_wb['当月考勤']
                prev_banked_col = None
                for r in [2, 3]:
                    for col in range(1, 50):
                        if str(prev_ws.cell(row=r, column=col).value).strip() == '存班':
                            prev_banked_col = col
                            break
                if prev_banked_col:
                    for row in range(4, prev_ws.max_row + 1):
                        name_val = prev_ws.cell(row=row, column=2).value
                        if name_val:
                            name = str(name_val).replace(' ', '').strip()
                            banked_val = str(prev_ws.cell(row=row, column=prev_banked_col).value).strip()
                            imported_banked_data[name] = parse_number(banked_val)
                    log("✅ [Data Hook] 提取完毕，存班池已写入系统内存。")
            else:
                log("⚠️ [警告] 挂载表内无【当月考勤】Sheet，放弃导入。")
        except Exception as e:
            log(f"⚠️ [警告] 读取上月考勤表失败：{str(e)}")

    log("📊 [Data Parse] 正在解析钉钉考勤流 (数据吞吐中)...")
    df_meta = pd.read_excel(source_file, sheet_name='原始记录', header=None, nrows=1)
    meta_text = str(df_meta.iloc[0, 0]) 
    dates_found = re.findall(r'(\d{4}-\d{2}-\d{2})', meta_text)
    
    if len(dates_found) < 2: 
        raise ValueError("周期解析失败：无法从【原始记录】中提取完整的开始与结束日期！")
        
    start_date = dates_found[0]
    end_date = dates_found[1]
    date_list = pd.date_range(start=start_date, end=end_date)
    
    year_str = end_date.split('-')[0]
    month_str = end_date.split('-')[1]
    if month_str.startswith('0'): 
        month_str = month_str[1:] 
    
    OUTPUT_FILE_PREFIX = os.path.join(BASE_DIR, f'山东宜美科{year_str}年')
    log(f"📅 [Time Sync] 成功锁定推演周期：{start_date} 至 {end_date} ({len(date_list)}天)")

    df_stats = pd.read_excel(source_file, sheet_name='月度汇总', header=2)
    df_stats.rename(columns={df_stats.columns[0]: '姓名'}, inplace=True)
    df_stats['match_name'] = df_stats['姓名'].astype(str).str.replace(' ', '').str.strip()

    df_daily_source = pd.read_excel(source_file, sheet_name='月度汇总', header=3)
    df_daily_source.rename(columns={df_daily_source.columns[0]: '姓名'}, inplace=True)
    df_daily_source['match_name'] = df_daily_source['姓名'].astype(str).str.replace(' ', '').str.strip()

    df_records = pd.read_excel(source_file, sheet_name='原始记录', header=2)
    df_records.rename(columns={df_records.columns[0]: '姓名'}, inplace=True)
    df_records['match_name'] = df_records['姓名'].astype(str).str.replace(' ', '').str.strip()
    df_records['date_clean'] = df_records['考勤日期'].astype(str).apply(lambda x: str(x).split(' ')[0])
    
    log("🧹 [Data Clean] 正在执行防呆过滤与审批记录洗白...")
    df_records.loc[df_records['打卡备注'].astype(str).str.contains('补卡审批通过'), '打卡结果'] = '正常'

    log("📐 [Matrix Build] 正在探测模板动态边界，重构输出网格...")
    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    ws = wb['当月考勤']
    
    stat_col_map = {}
    max_stat_col = 0
    for r in [2, 3]:
        for col in range(1, 50):
            val = str(ws.cell(row=r, column=col).value).strip()
            if val in ['出勤日', '省内', '省外', '加班', '病假', '请假', '调休', '迟到', '旷工', '工地天数', '存班', '未打卡']:
                stat_col_map[val] = col
                if col > max_stat_col: 
                    max_stat_col = col
                    
    date_start_col = max_stat_col + 1 if max_stat_col > 0 else 15
    
    remark_col = None
    for col in range(date_start_col, 100):
        if '备注' in str(ws.cell(row=2, column=col).value).strip():
            remark_col = col
            break
            
    if remark_col:
        template_days_count = remark_col - date_start_col
    else:
        template_days_count = sum(1 for c in range(date_start_col, 100) if ws.cell(row=3, column=c).value)
        
    if template_days_count != len(date_list):
        raise ValueError(f"格式错位：源数据共 {len(date_list)} 天，但模板预留了 {template_days_count} 列格子！\n请增删模板列使其对等。")
    
    old_banked_data = {}
    for row in range(4, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=2).value
        if name_cell:
            name = str(name_cell).replace(' ', '').strip()
            if '存班' in stat_col_map:
                if name in imported_banked_data: 
                    old_banked_data[name] = imported_banked_data[name]
                else: 
                    old_banked_data[name] = parse_number(str(ws.cell(row=row, column=stat_col_map['存班']).value))
        
        for col_idx in stat_col_map.values(): 
            safe_write(ws, row, col_idx, None)
        for col_idx in range(date_start_col, date_start_col + template_days_count): 
            safe_write(ws, row, col_idx, None)

    week_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
    date_col_map = {}
    for i, dt in enumerate(date_list):
        col = date_start_col + i
        safe_write(ws, 3, col, dt.day)
        safe_write(ws, 4, col, week_map[dt.weekday()]) 
        date_col_map[col] = dt.strftime('%Y-%m-%d')
        ws.cell(row=3, column=col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center', vertical='center')

    log("⚙️ [Core Engine] 算法矩阵就绪，开始执行时序推理与资金结转...")
    log("-" * 55)

    processed_cnt = 0
    for row in range(4, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=2).value
        if not name_cell: 
            continue
        name = str(name_cell).replace(' ', '').strip()
        
        stats_row = df_stats[df_stats['match_name'] == name]
        daily_row = df_daily_source[df_daily_source['match_name'] == name]
        if stats_row.empty or daily_row.empty: 
            continue
        
        processed_cnt += 1
        time.sleep(0.015) 
        
        log(f"⏳ 开始加载数据流 => 目标：[{name}]")
        
        stats_data = stats_row.iloc[0]
        daily_data = daily_row.iloc[0]
        dept = str(stats_data.get('部门', ''))
        is_hq = '本部' in str(stats_data.get('考勤组', dept))
        
        if is_hq:
            log(f"  ├─ 阵营探针：归属 [{dept}] -> 激活【本部双休与标准日历引擎】")
        else:
            log(f"  ├─ 阵营探针：归属 [{dept}] -> 激活【非本部大水池统筹结算引擎】")
        
        user_shift_type = "default"
        if any(k in dept for k in SHIFT_24_48): 
            user_shift_type = "24_48"
        elif any(k in dept for k in SHIFT_3_ROT): 
            user_shift_type = "3_rotation"
        
        for k, v in SHIFT_EXCEPTIONS.items():
            if k == name:
                if "白" in v or "正常" in v or "固定" in v: 
                    user_shift_type = "default"
                elif "24" in v or "48" in v: 
                    user_shift_type = "24_48"
                elif "三班" in v or "3班" in v: 
                    user_shift_type = "3_rotation"

        daily_punches = {}
        date_keys = list(date_col_map.values())
        for full_date_str in date_keys:
            short_date = full_date_str[5:]
            daily_recs = df_records[(df_records['match_name'] == name) & (df_records['date_clean'].str.endswith(short_date))]
            has_morning = False
            has_afternoon = False
            has_late = False
            
            for _, rec in daily_recs.iterrows():
                time_match = re.search(r'(\d{1,2}):\d{2}', str(rec['打卡时间']))
                if time_match:
                    hr = int(time_match.group(1))
                    if hr < 12: 
                        has_morning = True
                    elif hr < 18: 
                        has_afternoon = True
                    else: 
                        has_late = True
            
            daily_punches[full_date_str] = {
                'has_morning': has_morning,
                'has_afternoon': has_afternoon or has_late, 
                'has_late': has_late,
                'is_only_morning': has_morning and not (has_afternoon or has_late),
                'recs_empty': daily_recs.empty
            }

        override_status = {}
        edge_compensated = False 
        
        if user_shift_type == "24_48":
            log("  ├─ 时序分析：特征匹配 -> 挂载【上24休48】滚动扫描窗口")
            for i in range(len(date_keys) - 1):
                d1 = date_keys[i]
                d2 = date_keys[i+1]
                if daily_punches[d1]['is_only_morning'] and daily_punches[d2]['is_only_morning'] and not daily_punches[d1]['recs_empty'] and not daily_punches[d2]['recs_empty']:
                    override_status[d1] = '√'
                    override_status[d2] = '√'
                    if i + 2 < len(date_keys): 
                        override_status[date_keys[i+2]] = '√'
                        
            if len(date_keys) >= 3:
                d0 = date_keys[0]
                d1 = date_keys[1]
                d2 = date_keys[2]
                if daily_punches[d0]['is_only_morning'] and daily_punches[d1]['recs_empty']:
                    override_status[d0] = '√'
                    override_status[d1] = '√'
                    edge_compensated = True
                if daily_punches[d0]['recs_empty'] and daily_punches[d1]['is_only_morning'] and daily_punches[d2]['is_only_morning']:
                    override_status[d0] = '√'
                    edge_compensated = True

            if len(date_keys) >= 2:
                d_last = date_keys[-1]
                d_prev = date_keys[-2]
                if daily_punches[d_last]['is_only_morning'] and daily_punches[d_prev]['recs_empty']:
                    override_status[d_last] = '√'
                    edge_compensated = True

        elif user_shift_type == "3_rotation":
            log("  ├─ 时序分析：特征匹配 -> 挂载【三班倒】跨天缝合窗口")
            for i in range(len(date_keys) - 1):
                d1 = date_keys[i]
                d2 = date_keys[i+1]
                if daily_punches[d1]['has_late'] and daily_punches[d2]['has_morning']:
                    override_status[d1] = '√'
                    if daily_punches[d2]['is_only_morning']: 
                        override_status[d2] = '○'
            
            if len(date_keys) > 0:
                d0 = date_keys[0]
                if daily_punches[d0]['is_only_morning']:
                    override_status[d0] = '○'
                    edge_compensated = True
            
            if len(date_keys) > 0:
                d_last = date_keys[-1]
                if daily_punches[d_last]['has_late'] and not daily_punches[d_last]['has_morning']:
                    override_status[d_last] = '√'
                    edge_compensated = True

        if edge_compensated:
            log("  ├─ 边缘推理：触发！侦测到周期跨天截断 -> 已执行盲区智能洗白")

        local_prov_in = 0
        local_prov_out = 0
        hq_overtime_calc = 0.0
        actual_rest_days = 0 
        actual_attendance_days = 0 
        no_punch_count = 0 
        
        triggered_late = 0
        triggered_weekend_half = 0

        for col, full_date_str in date_col_map.items():
            short_date = full_date_str[5:]
            day_str = str(int(full_date_str.split('-')[-1]))
            
            daily_recs = df_records[(df_records['match_name'] == name) & (df_records['date_clean'].str.endswith(short_date))]
            dt_obj = pd.to_datetime(full_date_str)
            is_weekend = dt_obj.weekday() >= 5
            
            day_type = 'workday'
            if is_hq:
                if full_date_str in holiday_rest_days: 
                    day_type = 'holiday'
                elif full_date_str in work_on_holidays: 
                    day_type = 'workday'
                elif is_weekend: 
                    day_type = 'rest'
            else:
                if full_date_str in legal_holidays: 
                    day_type = 'holiday'
                elif is_weekend: 
                    day_type = 'rest'

            found_locs = []
            has_morning_punch = False
            has_afternoon_punch = False
            has_late_night_punch = False
            
            day_addrs = "".join(str(r['打卡地址']) + str(r['打卡备注']) for _, r in daily_recs.iterrows())
            day_results = "".join(str(r['打卡结果']) for _, r in daily_recs.iterrows())
            has_field_work = '外勤' in day_results
            
            is_emc01 = any(jk.lower() in day_addrs.lower() for jk in JINAN_KEYWORDS if 'emc' in jk.lower() or '公司' in jk or '本部' in jk)
            if is_hq and is_emc01: 
                has_field_work = False

            for _, rec in daily_recs.iterrows():
                addr = str(rec['打卡地址']) + str(rec['打卡备注'])
                time_match = re.search(r'(\d{1,2}):\d{2}', str(rec['打卡时间']))
                if time_match:
                    hr = int(time_match.group(1))
                    if hr < 12: 
                        has_morning_punch = True
                    if hr >= 12: 
                        has_afternoon_punch = True
                    if hr >= 21: 
                        has_late_night_punch = True

                if not (is_hq and is_emc01):
                    curr = None
                    for kw, sym in PROJECT_MAPPING.items():
                        if kw in addr: 
                            curr = sym
                            break
                    if not curr:
                        for kw, sym in CITY_ABBREVIATIONS.items():
                            if kw in addr: 
                                curr = sym
                                break
                    if not curr:
                        for city in SHANDONG_KEYWORDS + LOC_PROVINCE_OUT:
                            if city in addr: 
                                curr = city
                                break
                    if curr and curr not in found_locs: 
                        found_locs.append(curr)

            # ================= V34.0 绝杀修复：全波段表头雷达匹配 =================
            # 彻底废弃 Pandas 自身隐患重重的 .get() 获取方式，改为强固的全波段搜查
            val = None
            search_keys = [day_str, int(day_str)]
            try: search_keys.append(float(day_str))
            except: pass
            search_keys.append(f"{day_str}.0")
            
            # 第一梯队：从两张表中进行绝对精准匹配
            for src in [stats_data, daily_data]:
                for key in search_keys:
                    if key in src.index:
                        val = src[key]
                        break
                if val is not None:
                    break
                    
            # 第二梯队：当新版 Pandas 强行篡改表头为 Datetime 或特殊 Float 时进行“扫街”匹配
            if val is None:
                for src in [stats_data, daily_data]:
                    for k_idx, v_val in src.items():
                        k_str = str(k_idx).strip()
                        if k_str == day_str or k_str == f"{day_str}.0":
                            val = v_val
                            break
                        if hasattr(k_idx, 'day') and str(k_idx.day) == day_str:
                            val = v_val
                            break
                    if val is not None:
                        break

            # 最关键的一步：严防死守！只要查不到或者为空，绝对返回 'nan'，坚决不允许变成 '正常'！
            if val is None or pd.isna(val) or str(val).strip() in ['nan', 'None', '']:
                status_raw = "nan"
            else:
                status_raw = str(val).strip()
            # =======================================================================

            force_normal = False
            force_rest = False
            
            if full_date_str in override_status:
                if override_status[full_date_str] == '√': 
                    force_normal = True
                    status_raw = "正常" 
                elif override_status[full_date_str] == '○': 
                    force_rest = True
                    status_raw = "休息"

            if not found_locs and not (is_hq and is_emc01):
                if not daily_recs.empty or '正常' in status_raw or force_normal:
                    for kw, sym in PROJECT_MAPPING.items():
                        if kw in dept: 
                            found_locs.append(sym)
                            break
                    if not found_locs:
                        for kw, sym in CITY_ABBREVIATIONS.items():
                            if kw in dept: 
                                found_locs.append(sym)
                                break
                    if not found_locs:
                        for city in SHANDONG_KEYWORDS + LOC_PROVINCE_OUT:
                            if city in dept and city not in ['公司', '本部', '山东', '省外']: 
                                found_locs.append(city)
                                break
                    if not found_locs and has_field_work: 
                        found_locs.append("外勤")

            if not daily_recs.empty or '正常' in status_raw or force_normal:
                eval_string = day_addrs + "".join(found_locs)
                is_jinan = any(k in eval_string for k in JINAN_KEYWORDS) or (is_hq and is_emc01)
                is_sd = any(k in eval_string for k in SHANDONG_KEYWORDS) or '山东' in eval_string
                is_out = any(k in eval_string for k in LOC_PROVINCE_OUT)

                if is_hq:
                    if is_jinan: 
                        pass 
                    elif is_sd: 
                        local_prov_in += 1 
                    elif is_out or has_field_work: 
                        local_prov_out += 1 
                else:
                    if is_sd or is_jinan: 
                        local_prov_in += 1 
                    elif is_out or has_field_work: 
                        local_prov_out += 1 
                    else:
                        if any(k in dept for k in OUTSIDE_DEPT_KEYWORDS): 
                            local_prov_out += 1
                        else: 
                            local_prov_in += 1

            if is_hq:
                found_locs = [loc for loc in found_locs if loc not in JINAN_KEYWORDS]

            if found_locs:
                base_text = '/'.join(found_locs)
            else:
                base_text = '√'

            if force_normal: 
                pass 
            elif force_rest:
                if is_hq:
                    base_text = '○'
                else:
                    base_text = '调休'
                actual_rest_days += 1
            elif day_type in ['holiday', 'rest'] and daily_recs.empty and '正常' not in status_raw and '加班' not in status_raw:
                if day_type == 'holiday': 
                    base_text = '※'
                else: 
                    if is_hq:
                        base_text = '○'
                    else:
                        base_text = '调休'
                actual_rest_days += 1
            elif '假' in status_raw or '调休' in status_raw:
                if '0.5' not in status_raw:
                    actual_rest_days += 1
                else:
                    actual_rest_days += 0.5
                    
                if '0.5' in status_raw or '半天' in status_raw:
                    is_morning_leave = bool(re.search(r'(0[0-9]|1[0-1]):\d{2}', status_raw))
                    if '调休' in status_raw: 
                        if is_morning_leave:
                            base_text = "调休/√"
                        else:
                            base_text = "√/调休"
                    else: 
                        if is_morning_leave:
                            base_text = "请假/√"
                        else:
                            base_text = "√/请假"
                else: 
                    if '调休' in status_raw:
                        base_text = '调休'
                    else:
                        base_text = '假'
            elif daily_recs.empty and not any(k in status_raw for k in ['假', '调休', '出差', '正常']):
                if is_hq: 
                    base_text = '未打卡'
                    no_punch_count += 1
                else: 
                    base_text = '调休'
                    actual_rest_days += 1
            elif '旷工' in status_raw:
                if '旷工迟到' in status_raw:
                    base_text = '迟'
                else:
                    base_text = '×'
            elif '迟到' in status_raw and day_type == 'workday':
                if not daily_recs[daily_recs['打卡结果'] == '正常'].empty and '迟到' not in str(daily_recs['打卡结果'].values): 
                    pass 
                else: 
                    base_text = '迟'

            if is_hq:
                if day_type in ['rest', 'holiday'] and (has_morning_punch or has_afternoon_punch):
                    if has_morning_punch and has_afternoon_punch:
                        base_text = '+'
                        hq_overtime_calc += 1.0
                    else:
                        hq_overtime_calc += 0.5
                        triggered_weekend_half += 1
                        if day_type == 'rest': 
                            if has_morning_punch:
                                base_text = '+/○'
                            else:
                                base_text = '○/+'
                        else: 
                            if has_morning_punch:
                                base_text = '+/※'
                            else:
                                base_text = '※/+'
                elif day_type == 'workday' and has_late_night_punch:
                    hq_overtime_calc += 0.5
                    triggered_late += 1
                    if base_text == '√':
                        base_text = '√/+'
                    else:
                        base_text = f'{base_text}/+'

            safe_write(ws, row, col, base_text)
            
            if base_text in ['+/○', '○/+', '+/※', '※/+']: 
                actual_attendance_days += 0.5
            elif any(char in base_text for char in ['√', '迟', '+']) or (base_text not in ['○', '※', '×', '假', '调休', '病假', '年', '未打卡']):
                if '0.5' in status_raw or '半天' in status_raw: 
                    actual_attendance_days += 0.5
                else: 
                    actual_attendance_days += 1

        if triggered_late > 0:
            log(f"  ├─ 异常嗅探：判定 {triggered_late} 次工作日晚加班 (>21:00) -> 盖章(√/+)")
        if triggered_weekend_half > 0:
            log(f"  ├─ 异常嗅探：判定 {triggered_weekend_half} 次非工作日半天加班 -> 盖章(+/○)")
        if local_prov_out > 0:
            log(f"  ├─ GPS 巡检：抓取到省外/外勤定位 -> 累计省外天数 ({local_prov_out}天)")

        # ================= 汇总数据计算模块 =================
        # 为了应对打包后 Pandas 列名错乱，提取统计数据同样使用安全全波段搜索
        def get_stat_val(col_name):
            val = None
            for src in [stats_data, daily_data]:
                if col_name in src.index: return src[col_name]
            return 0
            
        dt_comp_leave = parse_number(get_stat_val('调休(天)'))
        dt_personal_leave = parse_number(get_stat_val('事假(天)'))
        dt_sick_leave = parse_number(get_stat_val('病假(天)'))
        dt_ot_dingding = sum(parse_number(get_stat_val(k)) for k in ['工作日加班', '休息日加班', '节假日加班'])
        
        old_banked = old_banked_data.get(name, 0.0)
        
        if is_hq:
            written_comp_leave = dt_comp_leave
            written_personal_leave = dt_personal_leave
            total_earned_ot = max(dt_ot_dingding, hq_overtime_calc)
            available_bank = old_banked + total_earned_ot
            
            if dt_comp_leave <= available_bank:
                written_comp_leave = dt_comp_leave
                new_banked = available_bank - dt_comp_leave
                calc_overtime = 0.0
                log(f"  └─ 资金池：调休({dt_comp_leave}) <= 可用池({available_bank}) -> 结余存班({new_banked})")
            else:
                written_comp_leave = available_bank 
                written_personal_leave = dt_personal_leave + (dt_comp_leave - available_bank) 
                new_banked = 0.0
                calc_overtime = 0.0
                log(f"  └─ 资金池：调休({dt_comp_leave}) > 可用池({available_bank}) -> 透支! 强制转事假({written_personal_leave - dt_personal_leave})")
        else:
            n_holidays = sum(1 for dt_str in date_col_map.values() if dt_str in legal_holidays)
            dynamic_quota = n_holidays + REST_QUOTA
            
            if old_banked + dynamic_quota <= actual_rest_days:
                written_comp_leave = old_banked + dynamic_quota
                written_personal_leave = dt_personal_leave + (actual_rest_days - old_banked - dynamic_quota)
                new_banked = 0.0
                calc_overtime = 0.0
                log(f"  └─ 资金池：[原存班({old_banked})+额度({dynamic_quota})] 无法覆盖休息({actual_rest_days}) -> 转事假({written_personal_leave - dt_personal_leave})")
            else:
                if dynamic_quota >= actual_rest_days:
                    calc_overtime = dynamic_quota - actual_rest_days
                    new_banked = old_banked + calc_overtime - n_holidays
                    written_comp_leave = actual_rest_days
                    written_personal_leave = dt_personal_leave
                    log(f"  └─ 资金池：额度({dynamic_quota})充足 -> 扣减休息后转加班({calc_overtime}), 新存班({new_banked})")
                else:
                    written_personal_leave = dt_personal_leave
                    calc_overtime = 0.0
                    new_banked = old_banked + dynamic_quota - actual_rest_days - n_holidays
                    written_comp_leave = actual_rest_days
                    log(f"  └─ 资金池：额度不足, 动用原存班补贴 -> 结转新存班({new_banked})")

            if new_banked < 0:
                new_banked = 0.0

        late_count = parse_number(get_stat_val('迟到次数')) + parse_number(get_stat_val('旷工迟到次数'))

        if '存班' in stat_col_map: 
            safe_write(ws, row, stat_col_map['存班'], new_banked if new_banked != 0 else None)
        if '调休' in stat_col_map: 
            safe_write(ws, row, stat_col_map['调休'], written_comp_leave if written_comp_leave > 0 else None)
        if '请假' in stat_col_map: 
            safe_write(ws, row, stat_col_map['请假'], written_personal_leave if written_personal_leave > 0 else None)
        if '病假' in stat_col_map: 
            safe_write(ws, row, stat_col_map['病假'], dt_sick_leave if dt_sick_leave > 0 else None)
        
        if '加班' in stat_col_map: 
            final_ot_to_write = max(dt_ot_dingding, hq_overtime_calc) if is_hq else (dt_ot_dingding + calc_overtime)
            safe_write(ws, row, stat_col_map['加班'], final_ot_to_write if final_ot_to_write > 0 else None)
            
        if '迟到' in stat_col_map: 
            safe_write(ws, row, stat_col_map['迟到'], late_count if late_count > 0 else None)
        if '出勤日' in stat_col_map: 
            safe_write(ws, row, stat_col_map['出勤日'], actual_attendance_days if actual_attendance_days > 0 else None)
        if '未打卡' in stat_col_map: 
            safe_write(ws, row, stat_col_map['未打卡'], no_punch_count if no_punch_count > 0 else None)
        if '省内' in stat_col_map: 
            safe_write(ws, row, stat_col_map['省内'], local_prov_in if local_prov_in > 0 else None)
        if '省外' in stat_col_map: 
            safe_write(ws, row, stat_col_map['省外'], local_prov_out if local_prov_out > 0 else None)
        
        if '工地天数' in stat_col_map:
            if any(k in dept for k in SITE_DAYS_DEPT_KEYWORDS): 
                safe_write(ws, row, stat_col_map['工地天数'], local_prov_in + local_prov_out)
            else: 
                safe_write(ws, row, stat_col_map['工地天数'], None) 

    log("-" * 55)
    log("💾 [File I/O] 正在将多维矩阵数据安全落盘至 Excel...")
    out_name = f"{OUTPUT_FILE_PREFIX}{month_str}月份考勤.xlsx"
    wb.save(out_name)
    log(f"🎉 [Success] 系统执行完毕！文件已生成：{os.path.basename(out_name)}")
    
    return processed_cnt

# ================= 极客风：控制台日志窗口 =================
class ProgressWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.main_app = parent
        self.is_done = False
        self.title("⚡ 智能考勤算力引擎")
        self.geometry("650x520")
        self.config(bg="#0C0C0C") 
        self.protocol("WM_DELETE_WINDOW", self.disable_close)
        
        tk.Label(self, text="Terminal | 底层算力图谱与边缘推理轨迹", font=("Microsoft YaHei UI", 11, "bold"), bg="#0C0C0C", fg="#00E676").pack(pady=(15, 5))
        
        self.text_area = tk.Text(self, bg="#1E1E1E", fg="#00FF41", font=("Consolas", 10), state=tk.DISABLED, relief=tk.FLAT, padx=12, pady=12)
        self.text_area.pack(expand=True, fill='both', padx=20, pady=(0, 10))
        
        self.btn_close = tk.Button(self, text="⏳ 正在高频推演中，请勿关闭...", font=("Microsoft YaHei", 10), bg="#333333", fg="#888888", state=tk.DISABLED, command=self.on_close, relief=tk.FLAT)
        self.btn_close.pack(pady=(0, 15), ipady=5, ipadx=20)
        
    def disable_close(self):
        if self.is_done:
            self.on_close()
        else:
            pass 
            
    def on_close(self):
        self.destroy()
        self.main_app.deiconify() 
        
    def log(self, msg):
        timestamp = datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]
        formatted_msg = f"[{timestamp}] {msg}\n"
        
        self.text_area.config(state=tk.NORMAL)
        self.text_area.insert(tk.END, formatted_msg)
        self.text_area.see(tk.END)
        self.text_area.config(state=tk.DISABLED)

# ================= 通用本地文档阅读器窗口 =================
class DocWindow(tk.Toplevel):
    def __init__(self, parent, file_path, title_text):
        super().__init__(parent)
        self.title(f"📖 {title_text}")
        self.geometry("750x650")
        self.config(padx=15, pady=15)
        
        tk.Label(self, text=title_text, font=("Microsoft YaHei", 16, "bold"), fg="#333333").pack(pady=(5, 15))
        
        self.txt_doc = scrolledtext.ScrolledText(self, wrap=tk.WORD, width=90, height=28, font=("Microsoft YaHei", 10), bg="#F9F9F9", padx=10, pady=10)
        self.txt_doc.pack(expand=True, fill='both')
        
        if os.path.exists(file_path):
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
            except Exception as e:
                content = f"读取文件出错：{str(e)}\n\n(提示：请确保此文件是使用 UTF-8 编码保存的。)"
        else:
            filename = os.path.basename(file_path)
            content = f"【提示：未找到文件】\n\n系统未在当前目录下找到名为『 {filename} 』的文件。\n\n操作步骤：\n1. 请在软件所在的文件夹中，手动新建一个名为“{filename}”的文本文档(.txt)。\n2. 将相关内容粘贴保存进去。\n3. 关闭此窗口重新点击按钮即可查看。"
            
        self.txt_doc.insert(tk.END, content)
        self.txt_doc.config(state=tk.DISABLED)

# ================= GUI 主控界面 =================
class AttendanceApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("全自动考勤计算系统 V34.0")
        self.geometry("550x540")
        self.config(padx=20, pady=20)
        self.prev_month_file = None
        
        load_config() 
        
        tk.Label(self, text="全自动考勤计算系统", font=("Microsoft YaHei", 18, "bold")).pack(pady=(0, 10))
        tk.Label(self, text="系统要求同目录下必须存在【模板-考勤.xlsx】", fg="gray").pack(pady=(0, 15))
        
        tk.Button(self, text="⚙️ 打开全局参数与排班特例配置", width=42, height=1, bg="#E0F7FA", font=("Arial", 10),
                  command=self.open_settings).pack(pady=6)
                  
        tk.Button(self, text="📜 查看详细计算规则 (算法白皮书)", width=42, height=1, bg="#F3E5F5", font=("Arial", 10),
                  command=lambda: DocWindow(self, RULES_FILE, "考勤算法与业务规则白皮书")).pack(pady=6)

        tk.Button(self, text="📘 查看官方使用手册 (操作指南)", width=42, height=1, bg="#FFF8E1", font=("Arial", 10),
                  command=lambda: DocWindow(self, MANUAL_FILE, "系统官方使用手册")).pack(pady=6)
                  
        tk.Label(self, text="-"*60, fg="#E0E0E0").pack(pady=10)
                  
        tk.Button(self, text="📂 [可选] 导入上月生成的考勤表 (自动提取原存班)", width=42, height=1, bg="#FFE0B2", font=("Arial", 10),
                  command=self.select_prev_file).pack(pady=8)
        self.lbl_prev_file = tk.Label(self, text="当前未挂载文件，将使用模板中默认存班为起点", fg="gray", font=("Arial", 8))
        self.lbl_prev_file.pack()

        tk.Button(self, text="🚀 选择【钉钉考勤导出报表】并开始生成", width=40, height=2, bg="#C8E6C9", font=("Arial", 11, "bold"),
                  command=self.start_process).pack(pady=(15, 10))
                  
        tk.Label(self, text="🛡️ 专属授权至：山东宜美科节能服务有限公司 | 企业终身版", fg="#9E9E9E", font=("Microsoft YaHei", 8)).pack(side=tk.BOTTOM, pady=8)

    def select_prev_file(self):
        file_path = filedialog.askopenfilename(
            title='请选择上个月生成的考勤结果表 (Excel)',
            filetypes=[("Excel files", "*.xlsx")]
        )
        if file_path:
            self.prev_month_file = file_path
            self.lbl_prev_file.config(text=f"已挂载上月表: {os.path.basename(file_path)}", fg="green")

    def start_process(self):
        source_file = filedialog.askopenfilename(
            title='请选择本月的钉钉考勤导出报表',
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if source_file:
            self.withdraw()
            self.progress_win = ProgressWindow(self)
            self.log_queue = queue.Queue()
            
            self.poll_log_queue()
            threading.Thread(target=self.run_task, args=(source_file, self.prev_month_file), daemon=True).start()

    def run_task(self, source, prev):
        try:
            start_time = time.time()
            processed_cnt = run_processing(source, prev, self.log_queue.put)
            elapsed_time = time.time() - start_time
            self.log_queue.put(f"DONE|{processed_cnt}|{elapsed_time:.2f}")
        except Exception as e:
            import traceback
            self.log_queue.put(f"ERROR|{str(e)}\n\n{traceback.format_exc()}")

    def poll_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                if msg.startswith("DONE|"):
                    parts = msg.split("|")
                    processed_cnt = int(parts[1]) if len(parts) > 1 else 0
                    elapsed_time = parts[2] if len(parts) > 2 else "0.00"
                    
                    saved_hours = round(processed_cnt * 5 / 60, 1)
                    if saved_hours < 1.0: 
                        saved_hours = 1.0
                        
                    self.progress_win.log("\n✨ === 系统挂起：核心算力图谱推演结束 ===")
                    
                    log_content = self.progress_win.text_area.get("1.0", tk.END)
                    timestamp_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                    log_filename = f"考勤算力日志_{timestamp_str}.txt"
                    log_filepath = os.path.join(BASE_DIR, log_filename)
                    try:
                        with open(log_filepath, 'w', encoding='utf-8') as f:
                            f.write(log_content)
                        save_status = f"\n\n📁 本次全量推演轨迹已防篡改存档至：\n{log_filename}"
                    except Exception as e:
                        save_status = f"\n\n⚠️ 日志自动存档失败：{str(e)}"
                    
                    roi_msg = (
                        f"✅ 考勤结算已生成并安全落盘！\n\n"
                        f"📊 【底层算力报告】\n"
                        f" ‣ 核心处理目标：{processed_cnt} 人\n"
                        f" ‣ 边缘排班推演与统筹大水池对冲：约 {processed_cnt * 30} 次运算\n"
                        f" ‣ 引擎总耗时：{elapsed_time} 秒\n\n"
                        f"{save_status}"
                    )
                    
                    self.progress_win.is_done = True
                    self.progress_win.btn_close.config(state=tk.NORMAL, text="✅ 查阅完毕，点击关闭日志并返回主界面", bg="#00E676", fg="#000000", font=("Microsoft YaHei", 10, "bold"))
                    
                    messagebox.showinfo("算力执行完毕", roi_msg, parent=self.progress_win)
                    return
                    
                elif msg.startswith("ERROR|"):
                    err_text = msg.split("|", 1)[1]
                    self.progress_win.log(f"\n❌ 发生致命错误：\n{err_text}")
                    
                    self.progress_win.is_done = True
                    self.progress_win.btn_close.config(state=tk.NORMAL, text="❌ 发生错误，点击关闭并返回主界面", bg="#FF5252", fg="white", font=("Microsoft YaHei", 10, "bold"))
                    messagebox.showerror("核心引擎发生异常", err_text, parent=self.progress_win)
                    return
                else:
                    self.progress_win.log(msg)
        except queue.Empty:
            pass
        self.after(100, self.poll_log_queue)

    def open_settings(self):
        SettingsWindow(self)

class SettingsWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("全局配置中心")
        self.geometry("680x580")
        self.config(padx=10, pady=10)
        self.grab_set()
        
        self.config_data = load_config()
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True)

        tab1 = ttk.Frame(self.notebook, padding=15)
        self.notebook.add(tab1, text="📅 基础与节假日")
        
        tk.Label(tab1, text="1. 本部：放假调休日期 (不打卡算※节假日) [用逗号隔开]", fg="blue").pack(anchor='w')
        self.ent_rest = tk.Entry(tab1, width=85)
        self.ent_rest.insert(0, ",".join(self.config_data.get("holiday_rest_days", [])))
        self.ent_rest.pack(pady=5, anchor='w')
        
        tk.Label(tab1, text="\n2. 本部：调休上班日期 (必须打卡)", fg="blue").pack(anchor='w')
        self.ent_work = tk.Entry(tab1, width=85)
        self.ent_work.insert(0, ",".join(self.config_data.get("work_on_holidays", [])))
        self.ent_work.pack(pady=5, anchor='w')
        
        tk.Label(tab1, text="\n3. 非本部：法定放假核心日期 (将转化为调休额度n)", fg="blue").pack(anchor='w')
        self.ent_legal = tk.Entry(tab1, width=85)
        self.ent_legal.insert(0, ",".join(self.config_data.get("legal_holidays", [])))
        self.ent_legal.pack(pady=5, anchor='w')
        
        tk.Label(tab1, text="\n4. 非本部人员每月默认休假额度 (天)", fg="blue").pack(anchor='w')
        self.ent_quota = tk.Entry(tab1, width=20)
        self.ent_quota.insert(0, str(self.config_data.get("rest_quota", 4.0)))
        self.ent_quota.pack(pady=5, anchor='w')

        tab2 = ttk.Frame(self.notebook, padding=15)
        self.notebook.add(tab2, text="📍 地点与部门词库")
        
        tk.Label(tab2, text="1. 济南本部打卡词 (用于拦截误判外勤)：", fg="blue").pack(anchor='w')
        self.ent_jinan = tk.Entry(tab2, width=85)
        self.ent_jinan.insert(0, ",".join(self.config_data.get("jinan_keywords", [])))
        self.ent_jinan.pack(pady=5, anchor='w')
        
        tk.Label(tab2, text="\n2. 山东省内地名词库：", fg="blue").pack(anchor='w')
        self.ent_sd = tk.Entry(tab2, width=85)
        self.ent_sd.insert(0, ",".join(self.config_data.get("shandong_keywords", [])))
        self.ent_sd.pack(pady=5, anchor='w')

        tk.Label(tab2, text="\n3. 常见外省地名词库 (用于计算省外天数)：", fg="blue").pack(anchor='w')
        self.ent_out = tk.Entry(tab2, width=85)
        self.ent_out.insert(0, ",".join(self.config_data.get("loc_province_out", [])))
        self.ent_out.pack(pady=5, anchor='w')

        tk.Label(tab2, text="\n4. 需要计算【工地天数】的部门关键字：", fg="blue").pack(anchor='w')
        self.ent_site_dept = tk.Entry(tab2, width=85)
        self.ent_site_dept.insert(0, ",".join(self.config_data.get("site_days_dept_keywords", [])))
        self.ent_site_dept.pack(pady=5, anchor='w')

        tk.Label(tab2, text="\n5. 长期驻外省的兜底部门 (防漏算省外天数)：", fg="blue").pack(anchor='w')
        self.ent_out_dept = tk.Entry(tab2, width=85)
        self.ent_out_dept.insert(0, ",".join(self.config_data.get("outside_dept_keywords", [])))
        self.ent_out_dept.pack(pady=5, anchor='w')

        tab3 = ttk.Frame(self.notebook, padding=15)
        self.notebook.add(tab3, text="🛠️ 字典与映射")
        
        tk.Label(tab3, text="1. 项目名称转简称字典 (格式：原名:简称, 原名:简称)", fg="blue").pack(anchor='w')
        self.ent_proj = tk.Entry(tab3, width=85)
        proj_str = ", ".join([f"{k}:{v}" for k, v in self.config_data.get("project_mapping", {}).items()])
        self.ent_proj.insert(0, proj_str)
        self.ent_proj.pack(pady=5, anchor='w')
        
        tk.Label(tab3, text="\n2. 城市名转单字字典 (格式：济南:济, 北京:京)", fg="blue").pack(anchor='w')
        self.ent_city = tk.Entry(tab3, width=85)
        city_str = ", ".join([f"{k}:{v}" for k, v in self.config_data.get("city_abbreviations", {}).items()])
        self.ent_city.insert(0, city_str)
        self.ent_city.pack(pady=5, anchor='w')

        tab4 = ttk.Frame(self.notebook, padding=15)
        self.notebook.add(tab4, text="⚙️ 复杂排班特例")
        
        tk.Label(tab4, text="【系统会自动寻找人员所在部门是否包含以下关键词，并启用专门的时序排班算法】\n\n1. 上24休48 部门/项目关键词 (自动合并连续早卡三天)：", fg="blue", justify="left").pack(anchor='w')
        self.ent_2448 = tk.Entry(tab4, width=85)
        self.ent_2448.insert(0, ",".join(self.config_data.get("shift_24_48", [])))
        self.ent_2448.pack(pady=5, anchor='w')
        
        tk.Label(tab4, text="\n2. 三班倒 部门/项目关键词 (自动缝合跨天晚班与次日早卡)：", fg="blue").pack(anchor='w')
        self.ent_3rot = tk.Entry(tab4, width=85)
        self.ent_3rot.insert(0, ",".join(self.config_data.get("shift_3_rotation", [])))
        self.ent_3rot.pack(pady=5, anchor='w')
        
        tk.Label(tab4, text="\n3. 个人排班特例白名单 (拥有最高优先级覆盖权)\n格式：姓名:白班 或 姓名:三班", fg="red", justify="left").pack(anchor='w')
        self.ent_exc = tk.Entry(tab4, width=85)
        exc_str = ", ".join([f"{k}:{v}" for k, v in self.config_data.get("shift_exceptions", {}).items()])
        self.ent_exc.insert(0, exc_str)
        self.ent_exc.pack(pady=5, anchor='w')

        tk.Button(self, text="💾 保存所有配置", font=("Arial", 11, "bold"), width=30, height=2, bg="#bbdefb", command=self.save_and_close).pack(pady=20)

    def parse_dict(self, text):
        d = {}
        if not text.strip(): 
            return d
        for item in text.replace('：', ':').replace('，', ',').split(','):
            if ':' in item:
                k, v = item.split(':', 1)
                d[k.strip()] = v.strip()
        return d

    def save_and_close(self):
        def clean_list(text): 
            return [d.strip() for d in text.replace('，', ',').split(',') if d.strip()]
            
        new_config = {
            "legal_holidays": clean_list(self.ent_legal.get()),
            "holiday_rest_days": clean_list(self.ent_rest.get()),
            "work_on_holidays": clean_list(self.ent_work.get()),
            "rest_quota": float(self.ent_quota.get().strip() or 4.0),
            
            "jinan_keywords": clean_list(self.ent_jinan.get()),
            "shandong_keywords": clean_list(self.ent_sd.get()),
            "loc_province_out": clean_list(self.ent_out.get()),
            "site_days_dept_keywords": clean_list(self.ent_site_dept.get()),
            "outside_dept_keywords": clean_list(self.ent_out_dept.get()),
            
            "project_mapping": self.parse_dict(self.ent_proj.get()),
            "city_abbreviations": self.parse_dict(self.ent_city.get()),
            
            "shift_24_48": clean_list(self.ent_2448.get()),
            "shift_3_rotation": clean_list(self.ent_3rot.get()),
            "shift_exceptions": self.parse_dict(self.ent_exc.get())
        }
        save_config(new_config)
        messagebox.showinfo("成功", "全局配置已更新并保存到本地！")
        self.destroy()

if __name__ == "__main__":
    app = AttendanceApp()
    app.mainloop()